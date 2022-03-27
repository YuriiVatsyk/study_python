import openpyxl

inv_file = openpyxl.load_workbook("inventory1.xlsx")
product_list = inv_file["inventory"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name) + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation total value per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = total_value_per_supplier.get(supplier_name) + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products with inventory under 10
    if inventory < 50:
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")
print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
